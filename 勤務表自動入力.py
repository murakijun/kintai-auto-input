#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
勤務表自動入力スクリプト  v3.0

Outlookの送信済みメール（始業時報告・終業時報告）を読み取り、
勤務表Excelへ始業時間・終業時間・勤務形態を自動入力します。

【使い方】
  1. config.ini を開いて sender_email（自分のメールアドレス）を設定する
  2. Outlookを起動した状態でスクリプトを実行する
       > python 勤務表自動入力.py
  3. 初回: pip install pywin32 openpyxl

【月の切り替えは不要】
  year / month が "auto" の場合は今月を自動で処理します。
  月ごとのExcelパスは自動で記憶されます。

【設定変更は config.ini で行います（このファイルは編集不要）】
"""

import re
import sys
import os
import configparser
from datetime import timedelta, date
from pathlib import Path


# ============================================================
# 固定定数（変更不要）
# ============================================================

CONFIG_FILE    = Path(__file__).parent / "config.ini"
SHEET_NAME     = "勤務表"
DATA_START_ROW = 6     # 1日のデータが入る行（A列が 4/1）

COL_START_TIME = 3     # C列: 始業時間
COL_END_TIME   = 4     # D列: 終業時間
COL_備考       = 16    # P列: 備考（勤務形態）

SUBJECT_START  = "始業時報告"
SUBJECT_END    = "終業時報告"

RE_START_TIME  = re.compile(r'始業時間[：:]\s*(\d{1,2}):(\d{2})')
RE_END_TIME    = re.compile(r'終業時間[：:]\s*(\d{1,2}):(\d{2})')
RE_WORK_TYPE   = re.compile(r'在宅[/／]出社勤務[：:]\s*(.+?)[\n\r]')

WORK_TYPE_MAP  = {
    "在宅勤務":       "在宅勤務(客先業務)",
    "出社勤務":       "",
    "在宅/出社勤務":  "",
    "在宅／出社勤務": "",
}


# ============================================================
# 設定ファイル
# ============================================================

def load_config() -> configparser.ConfigParser:
    """config.ini を読み込む。ファイルがなければデフォルト値で作成する。"""
    cfg = configparser.ConfigParser()

    # デフォルト値
    cfg["user"]         = {"sender_email": ""}
    cfg["target"]       = {"year": "auto", "month": "auto"}  # auto = 今年・今月
    cfg["excel"]        = {"path": ""}   # 後方互換のために残す
    cfg["excel_paths"]  = {}             # 月ごとのExcelパス {2026_04: path, ...}
    cfg["outlook"]      = {"folder_type": "5"}
    cfg["options"]      = {"preview_only": "false"}

    if CONFIG_FILE.exists():
        cfg.read(CONFIG_FILE, encoding="utf-8")
    else:
        save_config(cfg)
        print(f"[情報] config.ini を新規作成しました: {CONFIG_FILE}")

    return cfg


def save_config(cfg: configparser.ConfigParser) -> None:
    """設定をファイルに書き込む。"""
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        cfg.write(f)


# ============================================================
# Excelファイル選択
# ============================================================

def select_excel_via_dialog(initial_dir: str = None) -> str:
    """ファイル選択ダイアログを開いてExcelパスを返す。"""
    try:
        import tkinter as tk
        from tkinter import filedialog
    except ImportError:
        return None

    root = tk.Tk()
    root.withdraw()
    root.wm_attributes("-topmost", 1)   # ダイアログを最前面に表示

    path = filedialog.askopenfilename(
        parent=root,
        title="勤務表Excelファイルを選択してください",
        initialdir=initial_dir or str(Path.home() / "Downloads"),
        filetypes=[
            ("Excelファイル", "*.xlsx *.xlsm"),
            ("すべてのファイル", "*.*"),
        ],
    )
    root.destroy()
    return path if path else None


def resolve_excel_path(cfg: configparser.ConfigParser, year: int, month: int) -> str:
    """
    Excelファイルパスを決定する優先順位:
      1. config.ini の [excel_paths] に当月分が記録されていればそれを使用
      2. 旧形式 [excel] path（後方互換）
      3. ファイル選択ダイアログ（→ 当月分として config.ini に保存）
    """
    month_key = f"{year}_{month:02d}"   # 例: "2026_04"

    # ① 当月分のパスが記録済みか確認
    if not cfg.has_section("excel_paths"):
        cfg.add_section("excel_paths")

    path = cfg.get("excel_paths", month_key, fallback="").strip()

    if path and os.path.exists(path):
        return path

    if path and not os.path.exists(path):
        print(f"[警告] {year}年{month}月のExcelパスが見つかりません:\n  {path}")
        print("  ファイル選択ダイアログを開きます...\n")

    # ② 旧形式のパスを確認（後方互換）
    legacy = cfg.get("excel", "path", fallback="").strip()
    if legacy and os.path.exists(legacy):
        # 旧形式を当月分として移行保存
        cfg.set("excel_paths", month_key, legacy)
        save_config(cfg)
        return legacy

    # ③ ダイアログで選択
    print(f"  {year}年{month}月の勤務表Excelファイルを選択してください（ダイアログが開きます）")
    selected = select_excel_via_dialog()

    if not selected:
        print("[エラー] ファイルが選択されませんでした。")
        sys.exit(1)

    # 当月分として保存（翌月は別ファイルが自動的に求められる）
    cfg.set("excel_paths", month_key, selected)
    save_config(cfg)
    print(f"[情報] {year}年{month}月のパスを config.ini に保存しました。\n  {selected}\n")

    return selected


# ============================================================
# ユーティリティ
# ============================================================

def hhmm_to_timedelta(hour: int, minute: int) -> timedelta:
    return timedelta(hours=hour, minutes=minute)


def timedelta_to_str(td: timedelta) -> str:
    total = int(td.total_seconds())
    return f"{total // 3600:02d}:{(total % 3600) // 60:02d}"


# ============================================================
# メール解析
# ============================================================

def parse_start_email(body: str) -> dict:
    """始業時報告メール → 始業時間 + 当日の勤務形態"""
    result = {}

    m = RE_START_TIME.search(body)
    if m:
        result["start_time"] = hhmm_to_timedelta(int(m.group(1)), int(m.group(2)))

    m = RE_WORK_TYPE.search(body)
    if m:
        raw = m.group(1).strip()
        result["work_type"] = WORK_TYPE_MAP.get(raw, raw)

    return result


def parse_end_email(body: str) -> dict:
    """終業時報告メール → 終業時間（始業時間はバックアップ用）"""
    result = {}

    m = RE_END_TIME.search(body)
    if m:
        result["end_time"] = hhmm_to_timedelta(int(m.group(1)), int(m.group(2)))

    m = RE_START_TIME.search(body)
    if m:
        result["start_time_backup"] = hhmm_to_timedelta(int(m.group(1)), int(m.group(2)))

    return result


# ============================================================
# Outlook接続
# ============================================================

def get_my_email(namespace) -> str:
    """Outlookから現在のアカウントのメールアドレスを自動取得する。"""
    try:
        return namespace.CurrentUser.Address.lower()
    except Exception:
        return ""


def is_my_email(sender_addr: str, my_email: str) -> bool:
    """送信者アドレスが自分のものかどうかを判定する。"""
    if not my_email:
        return True   # 未設定の場合はフィルタなし
    return sender_addr.strip().lower() == my_email.strip().lower()


def get_emails_from_outlook(year: int, month: int,
                             folder_type: int, my_email: str) -> list:
    """
    Outlookの指定フォルダから対象月の勤務報告メールを取得する。
    送信者アドレスが my_email と一致するものだけを返す。
    """
    try:
        import win32com.client
    except ImportError:
        print("[エラー] pywin32がインストールされていません。")
        print("  > pip install pywin32  を実行してください。")
        sys.exit(1)

    print("Outlookに接続中...")
    outlook   = win32com.client.Dispatch("Outlook.Application")
    namespace = outlook.GetNamespace("MAPI")

    # メールアドレスが未設定の場合は Outlook から自動取得
    if not my_email:
        auto = get_my_email(namespace)
        if auto:
            my_email = auto
            print(f"  自分のアドレスを自動取得: {my_email}")
        else:
            print("  [警告] メールアドレスを特定できませんでした。"
                  "config.ini の sender_email を設定してください。")

    folder = namespace.GetDefaultFolder(folder_type)
    print(f"  対象フォルダ: {folder.Name}（{folder.Items.Count} 件）")

    emails      = []
    skipped_sender = 0

    for item in folder.Items:
        try:
            subject = item.Subject
            if SUBJECT_START not in subject and SUBJECT_END not in subject:
                continue

            sent = item.SentOn
            if sent.year != year or sent.month != month:
                continue

            # ★ 送信者フィルタ：自分のメールのみ処理
            sender_addr = getattr(item, "SenderEmailAddress", "")
            if not is_my_email(sender_addr, my_email):
                skipped_sender += 1
                continue

            emails.append({
                "date":    date(sent.year, sent.month, sent.day),
                "subject": subject,
                "body":    item.Body,
                "sender":  sender_addr,
                "type":    "start" if SUBJECT_START in subject else "end",
            })
        except Exception:
            continue

    if skipped_sender:
        print(f"  他の送信者のメールをスキップ: {skipped_sender} 件")

    emails.sort(key=lambda x: (x["date"], 0 if x["type"] == "start" else 1))
    return emails


# ============================================================
# Excel更新
# ============================================================

def update_excel(excel_path: str, work_data: dict, preview: bool = False) -> list:
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, keep_vba=True)

    if SHEET_NAME not in wb.sheetnames:
        print(f"[エラー] シート '{SHEET_NAME}' が見つかりません。")
        print(f"  利用可能なシート: {wb.sheetnames}")
        sys.exit(1)

    ws  = wb[SHEET_NAME]
    log = []

    for d in sorted(work_data.keys()):
        data      = work_data[d]
        row       = DATA_START_ROW + (d.day - 1)
        start_str = timedelta_to_str(data["start_time"]) if data.get("start_time") else "未取得"
        end_str   = timedelta_to_str(data["end_time"])   if data.get("end_time")   else "未取得"
        wtype     = data.get("work_type", None)
        wtype_str = wtype if wtype else "(空欄=出社)"

        log.append(
            f"  {d.month:02d}/{d.day:02d}"
            f"  始業:{start_str}"
            f"  終業:{end_str}"
            f"  勤務形態:{wtype_str}"
        )

        if not preview:
            if data.get("start_time") is not None:
                ws.cell(row=row, column=COL_START_TIME).value = data["start_time"]
            if data.get("end_time") is not None:
                ws.cell(row=row, column=COL_END_TIME).value = data["end_time"]
            if wtype is not None:
                ws.cell(row=row, column=COL_備考).value = wtype

    if not preview:
        wb.save(excel_path)

    return log


# ============================================================
# メイン処理
# ============================================================

def main():
    # --- 設定読み込み ---
    cfg         = load_config()
    folder_type = cfg.getint("outlook", "folder_type", fallback=5)
    preview     = cfg.getboolean("options", "preview_only", fallback=False)
    my_email    = cfg.get("user", "sender_email", fallback="").strip()

    # 年月の解決（"auto" または空白 → 今月を使用）
    today     = date.today()
    year_str  = cfg.get("target", "year",  fallback="auto").strip().lower()
    month_str = cfg.get("target", "month", fallback="auto").strip().lower()
    year      = today.year  if year_str  in ("auto", "") else int(year_str)
    month     = today.month if month_str in ("auto", "") else int(month_str)

    print("=" * 60)
    print(f"  勤務表自動入力  {year}年{month}月")
    if year_str == "auto" and month_str == "auto":
        print("  （年月: 今月を自動検出）")
    if preview:
        print("  ★ プレビューモード（Excelには書き込みません）")
    print("=" * 60)

    if not my_email:
        print("\n[注意] config.ini の sender_email が未設定です。")
        print("  Outlookから自動取得を試みます。\n")

    # --- Excelファイル決定 ---
    excel_path = resolve_excel_path(cfg, year, month)
    print(f"Excelファイル: {excel_path}\n")

    # --- メール取得 ---
    emails = get_emails_from_outlook(year, month, folder_type, my_email)

    start_mails = [e for e in emails if e["type"] == "start"]
    end_mails   = [e for e in emails if e["type"] == "end"]
    print(f"  取得: 始業時報告 {len(start_mails)} 件 / 終業時報告 {len(end_mails)} 件\n")

    if not emails:
        print("[警告] 対象期間のメールが見つかりませんでした。")
        print("  ・Outlookが起動しているか確認してください。")
        print("  ・config.ini の year / month が正しいか確認してください。")
        sys.exit(0)

    # --- メール解析 ---
    print("─" * 60)
    print("メール解析結果")
    print("─" * 60)

    work_data = {}

    def get_or_init(d):
        if d not in work_data:
            work_data[d] = {}
        return work_data[d]

    for email in start_mails:
        parsed = parse_start_email(email["body"])
        d      = email["date"]
        entry  = get_or_init(d)

        if parsed.get("start_time"):
            entry["start_time"] = parsed["start_time"]
        if "work_type" in parsed:
            entry["work_type"] = parsed["work_type"]

        st = timedelta_to_str(parsed["start_time"]) if parsed.get("start_time") else "?"
        wt = parsed.get("work_type", "?")
        print(f"  [始業] {d.month:02d}/{d.day:02d}  始業:{st}  勤務形態:{wt}")

    for email in end_mails:
        parsed = parse_end_email(email["body"])
        d      = email["date"]
        entry  = get_or_init(d)

        if parsed.get("end_time"):
            entry["end_time"] = parsed["end_time"]

        et   = timedelta_to_str(parsed["end_time"]) if parsed.get("end_time") else "?"
        note = ""

        if "start_time" not in entry and parsed.get("start_time_backup"):
            entry["start_time"] = parsed["start_time_backup"]
            note = "  ※始業時間を終業メールから補完"

        print(f"  [終業] {d.month:02d}/{d.day:02d}  終業:{et}{note}")

    # --- Excel書き込み ---
    print()
    print("─" * 60)
    print("【プレビュー】書き込み内容" if preview else "Excel書き込み内容")
    print("─" * 60)

    log = update_excel(excel_path, work_data, preview=preview)
    for line in log:
        print(line)

    print()
    if preview:
        print(f"プレビュー完了。{len(log)} 日分を確認しました。")
        print("（config.ini で preview_only = false にすると実際に書き込みます）")
    else:
        print(f"完了！{len(log)} 日分を書き込みました。")
        print(f"保存先: {excel_path}")


if __name__ == "__main__":
    main()
